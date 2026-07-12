# core/exporters.py
"""
Exportadores multi-formato para el Landing Stage.
Soporta CSV, Excel (.xlsx) y SQLite con modo append o replace.
Optimizados para procesamiento por lotes sin re-lectura de memoria completa.
"""

import os
import sqlite3
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def exportar(df: pd.DataFrame, destino: dict, perfil_nombre: str = "") -> int:
    """
    Punto de entrada unificado para volcar lotes al destino configurado.

    destino = {
        "tipo":           "csv" | "xlsx" | "sqlite",
        "carpeta":        "/ruta/landing/",
        "nombre_archivo": "cierre_tgm",   # sin extensión
        "modo":           "append" | "replace"
    }
    """
    tipo  = destino.get("tipo", "csv").lower()
    modo  = destino.get("modo", "append")
    carpeta = destino.get("carpeta", ".")
    nombre  = destino.get("nombre_archivo") or perfil_nombre or "exportacion"

    os.makedirs(carpeta, exist_ok=True)

    # Agrega columna de auditoría
    df = df.copy()
    df["_run_ts"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if tipo == "csv":
        return _a_csv(df, carpeta, nombre, modo)
    elif tipo == "xlsx":
        return _a_xlsx(df, carpeta, nombre, modo)
    elif tipo == "sqlite":
        return _a_sqlite(df, carpeta, nombre, modo)
    else:
        raise ValueError(f"Formato de destino no soportado: '{tipo}'")


# ── CSV ───────────────────────────────────────────────────────────────────────

def _a_csv(df, carpeta, nombre, modo):
    ruta = os.path.join(carpeta, f"{nombre}.csv")
    existe = os.path.exists(ruta)

    if modo == "replace" or not existe:
        df.to_csv(ruta, index=False, encoding="utf-8-sig")
    else:
        # append: agrega sin repetir encabezado ni alterar datos previos en RAM
        df.to_csv(ruta, mode="a", header=False, index=False, encoding="utf-8-sig")

    return len(df)


# ── XLSX OPTIMIZADO PARA MEMORIA (STREAM EN APPEND) ───────────────────────────

def _a_xlsx(df, carpeta, nombre, modo):
    ruta = os.path.join(carpeta, f"{nombre}.xlsx")
    existe = os.path.exists(ruta)

    if modo == "append" and existe:
        # Se abre con openpyxl directo para escribir en la última fila libre.
        # Evita usar pd.read_excel que consume toda la memoria cargando el histórico.
        wb = load_workbook(ruta)
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
    else:
        # Modo 'replace' o creación inicial del libro
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Guardamos provisionalmente para aplicar estilos base
        wb.save(ruta)
        _estilizar_xlsx(ruta)
        return len(df)

    wb.save(ruta)
    return len(df)


def _estilizar_xlsx(ruta):
    wb = load_workbook(ruta)
    ws = wb.active

    thin   = Side(style="thin", color="D0D8E4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Segoe UI")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[1].height = 30

    fill_par   = PatternFill(start_color="EDF3FB", end_color="EDF3FB", fill_type="solid")
    fill_impar = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = fill_par if row_idx % 2 == 0 else fill_impar
        for cell in row:
            cell.fill      = fill
            cell.alignment = left
            cell.border    = border
        ws.row_dimensions[row_idx].height = 20

    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)

    ws.freeze_panes = "A2"
    wb.save(ruta)


# ── SQLite ────────────────────────────────────────────────────────────────────

def _a_sqlite(df, carpeta, nombre, modo):
    ruta  = os.path.join(carpeta, f"{nombre}.db")
    tabla = _nombre_tabla(nombre)

    con = sqlite3.connect(ruta)
    try:
        if_exists = "append" if modo == "append" else "replace"
        # El motor SQL nativo gestiona eficientemente el append sin saturar memoria
        df.to_sql(tabla, con, if_exists=if_exists, index=False)
    finally:
        con.close()

    return len(df)


def _nombre_tabla(nombre: str) -> str:
    """Convierte un nombre de archivo en nombre válido de tabla SQLite."""
    import re
    return re.sub(r"[^a-zA-Z0-9_]", "_", nombre).lower() or "datos"